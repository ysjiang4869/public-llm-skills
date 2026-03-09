#!/usr/bin/env python3
"""
Write work records into the stats report Excel file.
Usage: python write_records.py <stats_report.xlsx> <records_json> [sheet_name]

records_json: JSON string or path to JSON file, array of:
  {"date": "2026-02-02", "req_type": "功能优化", "description": "...", "hours": 0.5}

Writes to columns: A=date, G=req_type, H=description, I=hours
"""
import sys
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from copy import copy

def write_records(stats_path, records, sheet_name="2月研发支撑"):
    wb = openpyxl.load_workbook(stats_path)
    ws = wb[sheet_name]

    header_cells = list(ws.iter_rows(min_row=1, max_row=1))[0]

    def copy_style(src, dst):
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)

    # Find first empty row
    start_row = 2
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            start_row = row[0].row
            break

    for idx, rec in enumerate(records):
        row_num = start_row + idx

        cell_a = ws.cell(row=row_num, column=1, value=rec["date"])
        copy_style(header_cells[0], cell_a)
        cell_a.font = Font(name='宋体', size=11)
        cell_a.alignment = Alignment(horizontal='center', vertical='center')

        cell_g = ws.cell(row=row_num, column=7, value=rec["req_type"])
        copy_style(header_cells[6], cell_g)
        cell_g.font = Font(name='宋体', size=11)
        cell_g.alignment = Alignment(horizontal='center', vertical='center')

        cell_h = ws.cell(row=row_num, column=8, value=rec["description"])
        copy_style(header_cells[7], cell_h)
        cell_h.font = Font(name='宋体', size=11)
        cell_h.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        cell_i = ws.cell(row=row_num, column=9, value=rec["hours"])
        copy_style(header_cells[8], cell_i)
        cell_i.font = Font(name='宋体', size=11)
        cell_i.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(stats_path)
    return {"written": len(records), "start_row": start_row}

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: write_records.py <stats_report.xlsx> <records_json_or_file> [sheet_name]")
        sys.exit(1)

    stats_path = sys.argv[1]
    records_arg = sys.argv[2]
    sheet = sys.argv[3] if len(sys.argv) > 3 else "2月研发支撑"

    try:
        records = json.loads(records_arg)
    except json.JSONDecodeError:
        with open(records_arg) as f:
            records = json.load(f)

    result = write_records(stats_path, records, sheet)
    print(json.dumps(result, ensure_ascii=False))
