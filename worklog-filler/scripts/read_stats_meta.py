#!/usr/bin/env python3
"""
Read metadata from the stats report: dropdown options and existing row count.
Usage: python read_stats_meta.py <stats_report.xlsx> [sheet_name]
Output: JSON to stdout
"""
import sys
import json
import openpyxl

def read_meta(stats_path, sheet_name="2月研发支撑"):
    wb = openpyxl.load_workbook(stats_path)
    ws = wb[sheet_name]

    req_types = []
    groups = []

    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and dv.formula1:
            options = [o.strip() for o in dv.formula1.strip('"').split(',') if o.strip()]
            # Detect which column the validation applies to by checking sqref
            sqref_str = str(dv.sqref)
            if 'G' in sqref_str:
                req_types = options
            elif 'L' in sqref_str:
                groups = options

    # Find first empty data row
    first_empty = 2
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        has_data = any(cell.value is not None for cell in row)
        if not has_data:
            first_empty = row[0].row
            break

    return {
        "sheet": sheet_name,
        "req_types": req_types,
        "groups": groups,
        "first_empty_row": first_empty
    }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: read_stats_meta.py <stats_report.xlsx> [sheet_name]")
        sys.exit(1)

    sheet = sys.argv[2] if len(sys.argv) > 2 else "2月研发支撑"
    meta = read_meta(sys.argv[1], sheet)
    print(json.dumps(meta, ensure_ascii=False, indent=2))
